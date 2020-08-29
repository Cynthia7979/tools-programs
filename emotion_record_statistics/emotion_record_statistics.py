"""
Notes on EmotionType values:

Positive and negative moods correspond to negative and positive integers.
The magnitude of the integer represents the "level of caution" AKA how bad it is.
Currently anxiety and depression are both -10, which is the lowest.

Notes on marks:

"好" - End of a month
"适中" - End of the record
"差" - (Vacant style used for debug)
"""
# import xlrd
import openpyxl as xl
import matplotlib
import numpy as np
import matplotlib.pyplot as plt
import logging

# workbook = xlrd.open_workbook(path)
# sheet = workbook.sheet_by_index(0)
# test_cell = sheet.cell(0,0)
# assert type(test_cell) is xlrd.sheet.Cell
# assert type(workbook) is xlrd.book.Book
# xf_index = test_cell.xf_index
# print(xf_index)
# xf_style = workbook.xf_list[xf_index]
# xf_background = xf_style.background
#
# fill_pattern = xf_background.fill_pattern
# pattern_colour_index = xf_background.pattern_colour_index
# background_colour_index = xf_background.background_colour_index
# print(background_colour_index)

PATH = 'emotion_record.xlsx'
PATH_ = 'D://情绪记录表.xlsx'
START_MONTH = 4
START_DATE = 16
LOGGING_LEVEL = logging.DEBUG


class EmotionType(object):
    def __init__(self, style_name, color_name, value: int):
        self.style_name = style_name
        self.emotion_name = color_name
        self.value = value

    def __repr__(self):
        return self.emotion_name

    def __hash__(self):
        return hash(self.emotion_name)+hash(self.style_name)+hash(self.value)


class Emotion(EmotionType):
    def __init__(self, emotion_type: EmotionType, time: str):
        """
        :param emotion_type: EmotionType object
        :param time: formatted string `hh:mm`
        """
        super().__init__(emotion_type.style_name, emotion_type.emotion_name, emotion_type.value)
        self.time, self.hour, self.minute = [time] + time.split(':')

    def __repr__(self):
        return f"({self.time}-{self.emotion_name})"

    def __eq__(self, other):
        assert isinstance(other, Emotion) or isinstance(other, EmotionType)
        return self.style_name==other.style_name and self.emotion_name==other.emotion_name and self.value==other.value

    def __hash__(self):
        return super().__hash__()


class Day(object):
    def __init__(self, date: int, *emotions):
        self.date = date
        self.emotions = list(emotions) if emotions else []
        self.emotions_with_no_value = 0
        self.notations = {}

    @property
    def avg_emotion(self):
        return sum((e.value for e in self.emotions)) / (len(self.emotions) - self.emotions_with_no_value)

    def add_emotion(self, emotion_type, time: str):
        assert isinstance(emotion_type, EmotionType)
        self.emotions.append(Emotion(emotion_type, time))
        if emotion_type.value == 0:
            self.emotions_with_no_value += 1

    def add_notation(self, *notation):
        for n in notation:
            if n in self.notations.keys():
                self.notations[n] += 1
            else:
                self.notations[n] = 1

    def __repr__(self):
        return f'Day {self.date} with average {self.avg_emotion} {"and notations" + str(self.notations) if self.notations else ""} '


class Month(object):  # It's too similar with Day!
    def __init__(self, month: int, *days):
        self.month = month
        self.days = list(days) if days else []
        self.days_with_no_value = 0

    @property
    def avg_emotion(self):
        return sum((day.avg_emotion for day in self.days)) / (len(self.days) - self.days_with_no_value)

    @property
    def notation_statistics(self):
        all_notations = {}
        for day in self.days:
            for k, v in day.notations.items():
                if k in all_notations.keys():
                    all_notations[k] += v
                else:
                    all_notations[k] = v
        return all_notations

    def add_day(self, day):
        assert isinstance(day, Day)
        self.days.append(day)
        if day.avg_emotion == 0:
            self.days_with_no_value += 1

    def __repr__(self):
        return f'Month {self.month} with days {tuple(self.days)} \n\t averaging {self.avg_emotion}, notations {self.notation_statistics}\n'


# Emotions (Individually declared to reuse later)
MISC_EMOTIONS = 'misc'
NUMB = EmotionType("着色 3", "numb", -2)
NONE = EmotionType("常规", '(none)', 0)

ANXIETY_FEAR = 'Anxiety & Fear'
ANXIOUS = EmotionType("着色 4", "anxious", -10)
NERVOUS = EmotionType("60% - 着色 4", "nervous", -5)

SADNESS = 'Sadness'
DEPRESSED = EmotionType("着色 1", "depressed", -10)
FRUSTRATED = EmotionType("60% - 着色 1", "frustrated", -5)
SAD = EmotionType("40% - 着色 1", "sad", -3)

HAPPINESS = 'Happiness'
ECSTASY = EmotionType("着色 6", 'Ecstasy', 10)
EXCITED = EmotionType("60% - 着色 6", 'excited', 5)
HAPPY = EmotionType("40% - 着色 6", 'happy', 3)
CHILLING_OUT = EmotionType("20% - 着色 6", "chilling out", 1)

ANGER = 'Anger'
RAGE = EmotionType("着色 2", 'rage', 7)
OFFENDED = EmotionType("60% - 着色 2", 'offended', 3)

MATCH_COLOR = {"着色 3": NUMB, "20% - 着色 6": CHILLING_OUT, "常规": NONE,
               "着色 4": ANXIOUS, "60% - 着色 4": NERVOUS,
               "着色 1": DEPRESSED, "60% - 着色 1": FRUSTRATED, "40% - 着色 1": SAD,
               "着色 6": ECSTASY, "60% - 着色 6": EXCITED, "40% - 着色 6": HAPPY,
               "着色 2": RAGE, "60% - 着色 2": OFFENDED}
TIME_PERIODS = [f"{h}:{m}" for h, m in zip(sorted(list(range(7, 23)) * 2), (00, 30) * 16)]
#                                       Every hour occurs twice        o'clocks and half hours
MATCH_NOTATION = {'F': 'Fatigue', 'P': 'Procrast', 'I': 'Ill', 'S': 'Social', 'A': 'Academic',
                  'E': 'Entertain', 'R': 'Fluct'}
CLASSIFY_EMOTIONS = {NUMB: MISC_EMOTIONS, NONE: MISC_EMOTIONS,
                     ANXIOUS: ANXIETY_FEAR, NERVOUS: ANXIETY_FEAR,
                     DEPRESSED: SADNESS, FRUSTRATED: SADNESS, SAD: SADNESS,
                     ECSTASY: HAPPINESS, EXCITED: HAPPINESS, HAPPY: HAPPINESS, CHILLING_OUT: HAPPINESS,
                     RAGE: ANGER, OFFENDED: ANGER}


def main(test=True, show_misc=True, show_happy=True):
    """
    :param test: Use the demo file
    :param show_misc: Show misc emotions
    """
    global LOGGER, months

    default_ch = logging.StreamHandler()
    default_ch.setLevel(LOGGING_LEVEL)
    LOGGER = logging.getLogger('Emotion_record')
    LOGGER.setLevel(logging.DEBUG)
    LOGGER.addHandler(default_ch)

    plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
    plt.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号

    workbook = xl.load_workbook(PATH) if test else xl.load_workbook(PATH_)
    sheet = workbook.active

    current_month_no = START_MONTH
    current_date_no = START_DATE
    current_month = Month(current_month_no)
    total_date_no = 0
    all_days = []
    months = []
    periods = []
    _last_day_period = False
    LOGGER.debug('  '.join((str(x) for x in range(36))))
    for i, row in enumerate(sheet.iter_rows(min_row=2, max_col=37)):  # All cells including period
        LOGGER.debug(str(i + 2) + ' '.join([c.style for c in row]))

        current_day = Day(current_date_no)

        for j, cell in enumerate(row[4:-1]):  # Cells in a day (row)
            try:
                current_day.add_emotion(MATCH_COLOR[cell.style], TIME_PERIODS[j])
            except KeyError as e:
                LOGGER.error(f'Unrecognized color at {(j, i)}: {cell.style}')
                raise e
            if cell.value:  # Notations
                for v in str(cell.value).split("|"):
                    try:
                        v = int(v)
                        for i in range(v):
                            current_day.add_emotion(MATCH_COLOR[cell.style], TIME_PERIODS[j])
                    except ValueError:
                        current_day.add_notation(v)
        current_date_no += 1
        total_date_no += 1
        current_month.add_day(current_day)
        all_days.append(current_day)

        if row[-1].style == '差':  # Had period
            if _last_day_period:
                periods[-1].append(total_date_no)
            else:
                periods.append([total_date_no])
                _last_day_period = True
        else:
            if _last_day_period:
                _last_day_period = False

        day_type = row[0].style
        if day_type == '好':  # Start of a month
            months.append(current_month)
            current_month_no += 1
            current_date_no = 1
            current_month = Month(current_month_no)
        elif day_type == '适中':  # End of the record
            months.append(current_month)
            break

    print(months)

    # -----------Start of Statistics-----------
    # 1. Notation bar graph
    _, ax1 = new_plot('Notation Statistics')
    n_ = months[0].notation_statistics
    x = list(range(len(n_)))
    heights = n_.values()
    keys = [MATCH_NOTATION[n] for n in n_.keys()]
    ax1.bar(x, heights, tick_label=keys, align='center')

    # 2. Daily Avg Emotion curve
    _, ax2 = new_plot('Average Daily Emotions')
    x_dates = [i for i in range(len(all_days))]
    y_avgs = [d.avg_emotion for d in all_days]
    ax2.plot(x_dates, y_avgs)
    for xy in zip(x_dates, y_avgs):
        ax2.annotate('%.2f' % xy[1], xy=xy)

    x_labels = [all_days[0].date]
    mth_offset = 1
    for x, d in enumerate(all_days[1:]):
        if d.date == 1:
            x_labels.append(f'{START_MONTH + mth_offset}月{d.date}')
            mth_offset += 1
            ax2.axvline(x+1, linestyle='--', color='gray')
        else:
            x_labels.append(d.date)
    plt.xticks([i for i in range(x_dates[0], x_dates[0] + len(x_dates))], labels=x_labels)

    ax2.axhline(0, color="r", linewidth=1)
    ax2.axhline(1, color="g", linewidth=1)

    highlight_period(ax2, periods)

    # 3. Semi-hourly emotion curves
    _, ax3 = new_plot('Hourly Emotions')
    emo_v_shourly = [[] for i in range(len(TIME_PERIODS))]  # Pre-calculate for 4.
    all_emo = [] # Pre-calculate for next ones
    for day in all_days:
        e_ = day.emotions
        all_emo.extend(e_)
        y_emvalues = [e.value for e in e_][:len(TIME_PERIODS)]
        ax3.plot(TIME_PERIODS, y_emvalues, color=np.random.rand(3, ))
        for i, e in enumerate(y_emvalues):  # Pre-calculate for the next one
            emo_v_shourly[i].append(e)

    # 4. Average Semi-hourly emotion curves
    _, ax4 = new_plot('Average Hourly Emotions')
    avg_emo_shourly = [sum(emo_v_shourly[i]) / len(emo_v_shourly[i]) for i in range(len(TIME_PERIODS))]
    ax4.axhline(0, color="r", linewidth=1)
    ax4.plot(TIME_PERIODS, avg_emo_shourly)

    # 5. Emotion types stacked bar graph, per day
    _, ax5 = new_plot('Emotion Types per Day')
    # Prepare to do statistics
    #                     MISC       ANGER     ANXIETY      SAD       HAPPY
    emotion_colormap = ['#A5A5A5', '#ED7D31', '#FFC000', '#4472C4', '#70AD47']
    if not show_misc: emotion_colormap.pop(0)
    if not show_happy: emotion_colormap.pop(-1)

    type_statistics_all = {MISC_EMOTIONS: 0, ANGER: 0, ANXIETY_FEAR: 0, SADNESS: 0, HAPPINESS: 0}
    if not show_misc: del type_statistics_all[MISC_EMOTIONS]
    if not show_happy: del type_statistics_all[HAPPINESS]

    for day_count, day in enumerate(all_days):
        type_statistics_day = {MISC_EMOTIONS: 0, ANGER: 0, ANXIETY_FEAR: 0, SADNESS: 0, HAPPINESS: 0}
        if not show_misc: del type_statistics_day[MISC_EMOTIONS]
        if not show_happy: del type_statistics_day[HAPPINESS]

        for e in day.emotions:
            if (show_misc or CLASSIFY_EMOTIONS[e] != MISC_EMOTIONS) and \
               (show_happy or CLASSIFY_EMOTIONS[e] != HAPPINESS) and \
               (e != NONE):
                type_statistics_day[CLASSIFY_EMOTIONS[e]] += 1
                type_statistics_all[CLASSIFY_EMOTIONS[e]] += 1

        current_bottom = 0
        for c, (k,v) in enumerate(type_statistics_day.items()):
            ax5.bar(day_count+1, v, bottom=current_bottom, color=emotion_colormap[c], label=k if c==0 else '')
            current_bottom += v
            if v!=0: ax5.annotate(str(v), (day_count+1, current_bottom-1))
    plt.xticks([i for i in range(x_dates[0]+1, x_dates[0]+len(x_dates))], labels=x_labels)
    highlight_period(ax5, periods, extra_offset=1)

    # 6. Emotion types pie graph, in total
    _, ax6 = new_plot('Emotion Types in Total')
    ax6.pie(list(type_statistics_all.values()),
            labels=[s for s in type_statistics_all.keys()],
            colors=emotion_colormap,
            explode=(0.1,) * len(list(type_statistics_all.values()))
            )
    pcoef = np.corrcoef(np.array([time_to_int(i) for i in TIME_PERIODS]), np.array(avg_emo_shourly))
    print(pcoef)

    plt.legend()
    plt.show()


def new_plot(title) -> (plt.Figure, plt.Axes):
    fig, ax = plt.subplots()
    fig.canvas.set_window_title(title)
    ax.set_title(title)
    return fig, ax


def time_to_int(time):
    return int(time.split(':')[0]) + {'30': 0.5, '0': 0, '00': 0}[time.split(':')[1]]


def highlight_period(ax, periods, extra_offset=0):
    for alpha_month, period in enumerate(periods):
        xmin, xmax = period[0], period[-1]
        ax.axvspan(xmin+extra_offset, xmax+extra_offset, color="#FF4F3B88")


if __name__ == '__main__':
    main(test=False, show_misc=True, show_happy=False)
